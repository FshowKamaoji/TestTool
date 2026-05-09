"""
スクリーンショットツール (メインモニタ専用・動的スケール補正・完全固定版)
- ドラッグで撮影範囲を指定（メインモニタのみ）
- F1キーで指定範囲をJPG保存（連番ファイル名）
- グローバルホットキー対応：本アプリがアクティブでなくてもF1で撮影可能
- スクリーンショットをExcelに縦順に自動貼り付け
    * 貼り付け先のシート、開始位置、画像間の空き行数を指定可能
    * 画像の直上に元ファイル名を自動で記録
必要ライブラリ: pip install Pillow pynput openpyxl
"""

import os
import sys

# ── 1. 最優先：Tkinterを読み込む前にWindowsのDPIスケーリングを無効化する ──
if sys.platform == "win32":
    try:
        import ctypes
        try:
            ctypes.windll.user32.SetProcessDpiAwarenessContext(-4)
        except Exception:
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except Exception:
                try:
                    ctypes.windll.shcore.SetProcessDpiAwareness(1)
                except Exception:
                    pass
    except Exception:
        pass

# DPI設定を済ませてからUIライブラリを読み込む
import tkinter as tk
from tkinter import filedialog, ttk
import math
import threading
from PIL import Image as PILImage, ImageGrab
from pynput import keyboard as pynput_keyboard

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── カラーパレット ──────────────────────────────────────────────────────────
BG       = "#1e1e2e"
PANEL    = "#2a2a3e"
ACCENT   = "#7c3aed"
ACCENT2  = "#a78bfa"
SUCCESS  = "#22c55e"
WARNING  = "#f59e0b"
TEXT     = "#e2e8f0"
SUBTEXT  = "#94a3b8"
BORDER   = "#3f3f5f"

_META_SHEET = "_scrn_meta"


# ── Windowsグローバルホットキー ────────────────────────────────────────────
class Win32GlobalHotkey:
    MOD_NONE  = 0x0000
    WM_HOTKEY = 0x0312
    WM_QUIT   = 0x0012
    VK_F1     = 0x70
    HOTKEY_ID = 0xC1A1

    def __init__(self, callback):
        self.callback     = callback
        self._thread      = None
        self._thread_id   = 0
        self._registered  = False
        self._success     = False
        self._ready_event = threading.Event()

    def start(self):
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        self._ready_event.wait(timeout=2.0)
        return self._success

    def _run(self):
        import ctypes
        from ctypes import wintypes
        user32   = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        self._thread_id = kernel32.GetCurrentThreadId()
        msg = wintypes.MSG()
        user32.PeekMessageW(ctypes.byref(msg), None, 0, 0, 0)
        if user32.RegisterHotKey(None, self.HOTKEY_ID, self.MOD_NONE, self.VK_F1):
            self._registered = True
            self._success    = True
        else:
            self._success = False
        self._ready_event.set()
        if not self._registered:
            return
        try:
            while True:
                ret = user32.GetMessageW(ctypes.byref(msg), None, 0, 0)
                if ret == 0 or ret == -1:
                    break
                if msg.message == self.WM_HOTKEY and msg.wParam == self.HOTKEY_ID:
                    try:
                        self.callback()
                    except Exception:
                        pass
                user32.TranslateMessage(ctypes.byref(msg))
                user32.DispatchMessageW(ctypes.byref(msg))
        finally:
            if self._registered:
                user32.UnregisterHotKey(None, self.HOTKEY_ID)
                self._registered = False

    def stop(self):
        if self._thread_id:
            try:
                import ctypes
                ctypes.windll.user32.PostThreadMessageW(self._thread_id, self.WM_QUIT, 0, 0)
            except Exception:
                pass
            self._thread_id = 0


class RegionSelector:
    def __init__(self, parent, callback):
        self.callback = callback
        self.start_x = 0
        self.start_y = 0
        self.rect_id = None
        
        # メインモニタのサイズを取得
        self.w = parent.winfo_screenwidth()
        self.h = parent.winfo_screenheight()

        self.win = tk.Toplevel(parent)
        self.win.title("")
        self.win.overrideredirect(True)
        
        # メインモニタの左上(+0+0)を起点にウィンドウを配置
        self.win.geometry(f"{self.w}x{self.h}+0+0")
        self.win.attributes("-alpha", 0.35)
        self.win.attributes("-topmost", True)
        self.win.configure(bg="#000010")
        self.win.focus_force()

        self.canvas = tk.Canvas(self.win, cursor="crosshair", bg="#000010", highlightthickness=0, borderwidth=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.create_text(
            self.w // 2, self.h // 2,
            text="ドラッグして撮影範囲を選択\n（Escキーでキャンセル）",
            fill="#ffffff", font=("Yu Gothic UI", 18, "bold"), justify="center", tags="hint",
        )
        self.canvas.bind("<ButtonPress-1>",   self._on_press)
        self.canvas.bind("<B1-Motion>",       self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.win.bind("<Escape>", self._cancel)

    def _on_press(self, event):
        self.start_x = event.x
        self.start_y = event.y

        self.canvas.delete("hint")
        if self.rect_id:
            self.canvas.delete(self.rect_id)
            self.rect_id = None

    def _on_drag(self, event):
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        
        self.rect_id = self.canvas.create_rectangle(
            self.start_x, self.start_y, event.x, event.y,
            outline="#a78bfa", width=2, fill="#7c3aed", stipple="gray25", tags="sel",
        )
        self.canvas.delete("size_label")
        w = abs(event.x - self.start_x)
        h = abs(event.y - self.start_y)
        self.canvas.create_text(
            event.x + 12, event.y + 12, text=f"{w} × {h}",
            fill="#ffffff", font=("Consolas", 11, "bold"), anchor="nw", tags="size_label",
        )

    def _on_release(self, event):
        region = (
            min(self.start_x, event.x),
            min(self.start_y, event.y),
            max(self.start_x, event.x),
            max(self.start_y, event.y),
        )
        self.win.destroy()
        
        if (region[2] - region[0]) < 5 or (region[3] - region[1]) < 5:
            self.callback(None)
        else:
            self.callback(region)

    def _cancel(self, _=None):
        self.win.destroy()
        self.callback(None)


class ScreenshotApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("スクリーンショットツール")
        self.root.geometry("480x640")
        self.root.resizable(False, False)
        self.root.configure(bg=BG)

        self.save_dir  = tk.StringVar(value=os.path.expanduser("~/Desktop"))
        self.region    = None
        self.counter   = 1
        self.f1_ready  = True
        self._listener     = None
        self._win32_hotkey = None
        self._hotkey_mode  = "none"
        self._capturing    = False

        self.excel_enabled       = tk.BooleanVar(value=True)
        self.excel_path          = tk.StringVar(value="")
        self.excel_start_row_var = tk.StringVar(value="1")
        self.excel_start_col_var = tk.StringVar(value="A")
        self.excel_sheet_var     = tk.StringVar(value="")
        self.excel_gap_var       = tk.StringVar(value="5")  # 画像間の空き行数
        self.excel_row           = 1
        self.excel_col           = 1
        self._suppress_trace     = False

        self._build_ui()
        self._install_excel_traces()
        self._start_hotkey()
        self._sync_counter()

    def _build_ui(self):
        root = self.root

        hdr = tk.Frame(root, bg=ACCENT, height=56)
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr, text="📸  スクリーンショットツール",
            bg=ACCENT, fg="#ffffff", font=("Yu Gothic UI", 14, "bold"),
        ).pack(side=tk.LEFT, padx=16, pady=12)

        body = tk.Frame(root, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)

        self._section(body, "① 保存先フォルダ")
        row1 = tk.Frame(body, bg=BG)
        row1.pack(fill=tk.X, pady=(4, 0))
        tk.Entry(
            row1, textvariable=self.save_dir,
            bg=PANEL, fg=TEXT, insertbackground=TEXT,
            relief=tk.FLAT, font=("Consolas", 10),
            highlightbackground=BORDER, highlightthickness=1,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6, padx=(0, 8))
        self._btn(row1, "参照…", self._select_folder, small=True).pack(side=tk.LEFT)

        self._section(body, "② 撮影範囲の指定")
        self.region_label = tk.Label(
            body, text="未設定", bg=PANEL, fg=SUBTEXT,
            font=("Consolas", 10), relief=tk.FLAT, anchor="w", padx=10, pady=6,
            highlightbackground=BORDER, highlightthickness=1,
        )
        self.region_label.pack(fill=tk.X, pady=(4, 0))
        self._btn(body, "ドラッグで範囲を選択", self._start_region_select).pack(fill=tk.X, pady=(8, 0))

        self._section(body, "③ スクリーンショット")
        info = tk.Frame(body, bg=PANEL, relief=tk.FLAT, highlightbackground=BORDER, highlightthickness=1)
        info.pack(fill=tk.X, pady=(4, 0))
        tk.Label(
            info, text="F1 キーで撮影・保存（他アプリ使用中でもOK）",
            bg=PANEL, fg=ACCENT2, font=("Yu Gothic UI", 11, "bold"), pady=10,
        ).pack()

        self._section(body, "④ Excelへの貼り付け")

        if not OPENPYXL_AVAILABLE:
            tk.Label(
                body,
                text="⚠ openpyxl が未インストールです\n  pip install openpyxl を実行してください",
                bg=PANEL, fg=WARNING, font=("Yu Gothic UI", 9), anchor="w",
                padx=10, pady=6, justify="left",
                highlightbackground=BORDER, highlightthickness=1,
            ).pack(fill=tk.X, pady=(4, 0))
        else:
            self._build_excel_ui(body)

        self.status_var = tk.StringVar(value="準備完了")
        self.status_bar = tk.Label(
            root, textvariable=self.status_var,
            bg=PANEL, fg=SUBTEXT, font=("Yu Gothic UI", 9),
            anchor="w", padx=12, pady=5, relief=tk.FLAT,
        )
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _build_excel_ui(self, body):
        tk.Checkbutton(
            body, text="スクリーンショットをExcelに縦順に貼り付ける",
            variable=self.excel_enabled,
            bg=BG, fg=TEXT, selectcolor=PANEL,
            activebackground=BG, activeforeground=TEXT,
            font=("Yu Gothic UI", 10), anchor="w",
            command=self._on_excel_toggle,
        ).pack(fill=tk.X, pady=(4, 0))

        row_f = tk.Frame(body, bg=BG)
        row_f.pack(fill=tk.X, pady=(4, 0))
        tk.Label(row_f, text="ファイル:", bg=BG, fg=SUBTEXT,
                 font=("Yu Gothic UI", 9), width=7, anchor="w").pack(side=tk.LEFT)
        tk.Entry(
            row_f, textvariable=self.excel_path,
            bg=PANEL, fg=TEXT, insertbackground=TEXT,
            relief=tk.FLAT, font=("Consolas", 9),
            highlightbackground=BORDER, highlightthickness=1,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5, padx=(0, 6))
        self._btn(row_f, "参照", self._select_excel, small=True).pack(side=tk.LEFT, padx=(0, 4))
        self._btn(row_f, "新規", self._new_excel,    small=True).pack(side=tk.LEFT)

        row_s = tk.Frame(body, bg=BG)
        row_s.pack(fill=tk.X, pady=(4, 0))
        tk.Label(row_s, text="シート:", bg=BG, fg=SUBTEXT,
                 font=("Yu Gothic UI", 9), width=7, anchor="w").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(
            row_s, textvariable=self.excel_sheet_var,
            font=("Yu Gothic UI", 9), state="normal",
        )
        self.sheet_combo["values"] = []
        self.sheet_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)
        tk.Label(row_s, text="（空欄=先頭シート / 存在しない=新規）",
                 bg=BG, fg=SUBTEXT, font=("Yu Gothic UI", 8)).pack(side=tk.LEFT, padx=(6, 0))

        row_p = tk.Frame(body, bg=BG)
        row_p.pack(fill=tk.X, pady=(4, 0))
        tk.Label(row_p, text="開始位置:", bg=BG, fg=SUBTEXT,
                 font=("Yu Gothic UI", 9), width=7, anchor="w").pack(side=tk.LEFT)
        tk.Label(row_p, text="列", bg=BG, fg=TEXT,
                 font=("Yu Gothic UI", 9)).pack(side=tk.LEFT)
        tk.Entry(
            row_p, textvariable=self.excel_start_col_var,
            width=5,
            bg=PANEL, fg=TEXT, insertbackground=TEXT,
            relief=tk.FLAT, font=("Consolas", 10),
            highlightbackground=BORDER, highlightthickness=1,
        ).pack(side=tk.LEFT, padx=(2, 10), ipady=3)
        tk.Label(row_p, text="行", bg=BG, fg=TEXT,
                 font=("Yu Gothic UI", 9)).pack(side=tk.LEFT)
        tk.Spinbox(
            row_p, textvariable=self.excel_start_row_var,
            from_=1, to=9999, width=5,
            bg=PANEL, fg=TEXT, insertbackground=TEXT,
            relief=tk.FLAT, font=("Consolas", 10),
            highlightbackground=BORDER, highlightthickness=1,
            buttonbackground=PANEL,
        ).pack(side=tk.LEFT, padx=(2, 6))
        tk.Label(row_p, text="（例: A, B, 1, 2）",
                 bg=BG, fg=SUBTEXT, font=("Yu Gothic UI", 8)).pack(side=tk.LEFT)

        row_g = tk.Frame(body, bg=BG)
        row_g.pack(fill=tk.X, pady=(4, 0))
        tk.Label(row_g, text="画像間隔:", bg=BG, fg=SUBTEXT,
                 font=("Yu Gothic UI", 9), width=7, anchor="w").pack(side=tk.LEFT)
        tk.Spinbox(
            row_g, textvariable=self.excel_gap_var,
            from_=0, to=99, width=5,
            bg=PANEL, fg=TEXT, insertbackground=TEXT,
            relief=tk.FLAT, font=("Consolas", 10),
            highlightbackground=BORDER, highlightthickness=1,
            buttonbackground=PANEL,
        ).pack(side=tk.LEFT, padx=(2, 6))
        tk.Label(row_g, text="行空ける（0で間隔なし）", bg=BG, fg=TEXT,
                 font=("Yu Gothic UI", 9)).pack(side=tk.LEFT)

        row_c = tk.Frame(body, bg=BG)
        row_c.pack(fill=tk.X, pady=(6, 0))
        self.excel_status_var = tk.StringVar(value="次の貼り付け位置: 列 A 行 1")
        tk.Label(
            row_c, textvariable=self.excel_status_var,
            bg=BG, fg=SUBTEXT, font=("Yu Gothic UI", 9), anchor="w",
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._btn(row_c, "位置リセット", self._reset_excel_row, small=True).pack(side=tk.LEFT)

    def _section(self, parent, text):
        tk.Label(
            parent, text=text, bg=BG, fg=SUBTEXT,
            font=("Yu Gothic UI", 9, "bold"), anchor="w",
        ).pack(fill=tk.X, pady=(14, 0))

    def _btn(self, parent, text, cmd, small=False):
        font = ("Yu Gothic UI", 9, "bold") if small else ("Yu Gothic UI", 10, "bold")
        return tk.Button(
            parent, text=text, command=cmd,
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT2, activeforeground="#ffffff",
            relief=tk.FLAT, cursor="hand2",
            font=font, padx=10 if small else 0, pady=4 if small else 8, borderwidth=0,
        )

    def _select_folder(self):
        path = filedialog.askdirectory(initialdir=self.save_dir.get())
        if path:
            self.save_dir.set(path)
            self._sync_counter()
            self._set_status(f"保存先: {path}", SUCCESS)

    def _start_region_select(self):
        self.root.withdraw()
        self.root.after(250, lambda: RegionSelector(self.root, self._on_region_selected))

    def _on_region_selected(self, region):
        self.root.deiconify()
        self.root.lift()
        if region:
            self.region = region
            x1, y1, x2, y2 = region
            self.region_label.config(
                text=f"({x1}, {y1})  →  ({x2}, {y2})   サイズ: {x2-x1} × {y2-y1} px",
                fg=TEXT,
            )
            self._set_status("撮影範囲を設定しました。F1キーで撮影できます。", SUCCESS)
        else:
            self._set_status("範囲の選択がキャンセルされました", WARNING)

    def _on_excel_toggle(self):
        if self.excel_enabled.get() and not self.excel_path.get():
            self._set_status("⚠ Excelファイルを選択または新規作成してください", WARNING)

    def _select_excel(self):
        path = filedialog.askopenfilename(
            initialdir=self.save_dir.get(),
            title="貼り付け先のExcelファイルを選択",
            filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
        )
        if path:
            self.excel_path.set(path)
            self._populate_sheet_list(path)
            self._load_excel_state(path)
            self._set_status(f"Excel: {os.path.basename(path)}  行 {self.excel_row}・列 {get_column_letter(self.excel_col)} から貼り付け", SUCCESS)

    def _new_excel(self):
        path = filedialog.asksaveasfilename(
            initialdir=self.save_dir.get(),
            title="新規Excelファイルの保存先",
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx")],
        )
        if path:
            try:
                wb = Workbook()
                wb.save(path)
                self.excel_path.set(path)
                self._populate_sheet_list(path)
                self._apply_start_position()
                self._update_excel_status()
                self._set_status(f"新規Excel作成: {os.path.basename(path)}", SUCCESS)
            except Exception as e:
                self._set_status(f"⚠ Excel作成エラー: {e}", WARNING)

    def _populate_sheet_list(self, path):
        if not hasattr(self, "sheet_combo"):
            return
        try:
            wb = load_workbook(path, read_only=True)
            names = [s for s in wb.sheetnames if s != _META_SHEET]
            wb.close()
            self.sheet_combo["values"] = names
        except Exception:
            self.sheet_combo["values"] = []

    def _parse_col(self, col_str):
        s = col_str.strip().upper()
        if not s:
            return 1
        if s.isdigit():
            return max(1, int(s))
        try:
            return column_index_from_string(s)
        except Exception:
            return 1

    def _first_data_sheet_name(self, wb):
        for s in wb.sheetnames:
            if s != _META_SHEET:
                return s
        return wb.sheetnames[0] if wb.sheetnames else "Sheet1"

    def _read_meta_for_sheet(self, wb, sheet_name):
        if _META_SHEET not in wb.sheetnames:
            return False
        meta = wb[_META_SHEET]

        for row_idx in range(1, meta.max_row + 1):
            v0 = meta.cell(row=row_idx, column=1).value
            v1 = meta.cell(row=row_idx, column=2).value
            v2 = meta.cell(row=row_idx, column=3).value
            if isinstance(v0, str) and v0 == sheet_name \
                    and isinstance(v1, int) and isinstance(v2, int):
                self.excel_row = max(1, v1)
                self.excel_col = max(1, v2)
                self._sync_ui_from_internal()
                return True

        a1 = meta["A1"].value
        b1 = meta["B1"].value
        if isinstance(a1, int) and isinstance(b1, int):
            self.excel_row = max(1, a1)
            self.excel_col = max(1, b1)
            self._sync_ui_from_internal()
            return True

        return False

    def _load_excel_state(self, path):
        sheet_name = self.excel_sheet_var.get().strip()
        try:
            wb = load_workbook(path)
            target = sheet_name or self._first_data_sheet_name(wb)
            loaded = self._read_meta_for_sheet(wb, target)
            wb.close()
            if not loaded:
                self._apply_start_position()
            self._update_excel_status()
        except Exception:
            self._apply_start_position()

    def _apply_start_position(self):
        try:
            self.excel_row = max(1, int(self.excel_start_row_var.get()))
        except (ValueError, TypeError):
            self.excel_row = 1
        self.excel_col = self._parse_col(self.excel_start_col_var.get())
        self._update_excel_status()

    def _sync_ui_from_internal(self):
        self._suppress_trace = True
        try:
            self.excel_start_row_var.set(str(self.excel_row))
            self.excel_start_col_var.set(get_column_letter(self.excel_col))
        finally:
            self._suppress_trace = False

    def _save_excel_state(self, wb, sheet_name):
        if _META_SHEET not in wb.sheetnames:
            meta = wb.create_sheet(_META_SHEET)
            meta.cell(row=1, column=1, value="sheet")
            meta.cell(row=1, column=2, value="row")
            meta.cell(row=1, column=3, value="col")
        else:
            meta = wb[_META_SHEET]
            if not (isinstance(meta.cell(row=1, column=1).value, str)
                    and meta.cell(row=1, column=1).value == "sheet"):
                meta.delete_rows(1, meta.max_row)
                meta.cell(row=1, column=1, value="sheet")
                meta.cell(row=1, column=2, value="row")
                meta.cell(row=1, column=3, value="col")

        target_row = None
        for r in range(2, meta.max_row + 1):
            if meta.cell(row=r, column=1).value == sheet_name:
                target_row = r
                break
        if target_row is None:
            target_row = meta.max_row + 1 if meta.max_row >= 1 else 2

        meta.cell(row=target_row, column=1, value=sheet_name)
        meta.cell(row=target_row, column=2, value=self.excel_row)
        meta.cell(row=target_row, column=3, value=self.excel_col)
        meta.sheet_state = "hidden"

    def _reset_excel_row(self):
        self._suppress_trace = True
        try:
            self.excel_start_row_var.set("1")
            self.excel_start_col_var.set("A")
        finally:
            self._suppress_trace = False
        self.excel_row = 1
        self.excel_col = 1
        self._update_excel_status()
        self._set_status("Excel貼り付け位置を A1（行 1・列 A）に初期化しました", SUCCESS)

    def _install_excel_traces(self):
        if not OPENPYXL_AVAILABLE:
            return
        self.excel_start_row_var.trace_add("write", self._on_start_pos_changed)
        self.excel_start_col_var.trace_add("write", self._on_start_pos_changed)
        self.excel_sheet_var.trace_add("write", self._on_sheet_changed)

    def _on_start_pos_changed(self, *_args):
        if self._suppress_trace:
            return
        self._apply_start_position()

    def _on_sheet_changed(self, *_args):
        if self._suppress_trace:
            return
        path = self.excel_path.get()
        sheet_name = self.excel_sheet_var.get().strip()
        if not path or not os.path.exists(path):
            self._apply_start_position()
            return
        try:
            wb = load_workbook(path)
            target = sheet_name or self._first_data_sheet_name(wb)
            loaded = self._read_meta_for_sheet(wb, target)
            wb.close()
            if not loaded:
                self._apply_start_position()
            self._update_excel_status()
        except Exception:
            self._apply_start_position()

    def _update_excel_status(self):
        if hasattr(self, "excel_status_var"):
            col_letter = get_column_letter(self.excel_col) if OPENPYXL_AVAILABLE else "A"
            self.excel_status_var.set(f"次の貼り付け位置: 列 {col_letter} 行 {self.excel_row} ")

    def _get_or_create_sheet(self, wb, sheet_name):
        if not sheet_name:
            ws = wb.active
            if ws.title == _META_SHEET:
                others = [s for s in wb.sheetnames if s != _META_SHEET]
                ws = wb[others[0]] if others else wb.create_sheet("Sheet1")
            return ws
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.create_sheet(sheet_name)

    def _paste_to_excel(self, filepath, logical_width_px, logical_height_px):
        if not self.excel_enabled.get():
            return
        if not OPENPYXL_AVAILABLE:
            return

        xl_path = self.excel_path.get()
        if not xl_path:
            self._set_status("⚠ Excelファイルを指定してください", WARNING)
            return

        if self.excel_row < 1 or self.excel_col < 1:
            self._apply_start_position()

        sheet_name = self.excel_sheet_var.get().strip()

        try:
            # 安全に空き行数(gap)を取得
            try:
                gap = int(self.excel_gap_var.get())
                gap = max(0, gap)
            except ValueError:
                gap = 1

            if os.path.exists(xl_path):
                wb = load_workbook(xl_path)
            else:
                wb = Workbook()

            ws = self._get_or_create_sheet(wb, sheet_name)
            col_letter = get_column_letter(self.excel_col)
            
            # 直上のセルにファイル名テキストを入力
            filename = os.path.basename(filepath)
            ws.cell(row=self.excel_row, column=self.excel_col, value=filename)

            # 画像はその「1行下」から貼り付ける
            img_start_row = self.excel_row + 1
            anchor = f"{col_letter}{img_start_row}"

            xl_img = XLImage(filepath)
            
            # 画像のExcel上での表示サイズを、ドラッグした見た目(論理ピクセル)に強制ロックする
            xl_img.width = logical_width_px
            xl_img.height = logical_height_px
            
            xl_img.anchor = anchor
            ws.add_image(xl_img)

            # 【根本解決】Excelの環境依存(DPIやフォント)による行の高さのブレを防ぐため、
            # プログラム側で強制的に行の高さを18.0ポイント(24px)に固定する。
            FIXED_ROW_HEIGHT_PT = 18.0
            
            # 画像の高さ(ポイント)
            img_height_pt = logical_height_px * 0.75
            
            # 画像が消費する物理行数を計算（切り上げ）
            rows_consumed = max(1, math.ceil(img_height_pt / FIXED_ROW_HEIGHT_PT))
            
            # ファイル名の行、画像が配置される行、および空き行(gap)の高さをすべて固定
            total_rows_to_format = 1 + rows_consumed + gap
            for i in range(total_rows_to_format):
                r = self.excel_row + i
                ws.row_dimensions[r].height = FIXED_ROW_HEIGHT_PT

            # セルの列幅も画像サイズに合わせて調整
            col_width = logical_width_px / 7.0
            if ws.column_dimensions[col_letter].width is None \
                    or ws.column_dimensions[col_letter].width < col_width:
                ws.column_dimensions[col_letter].width = col_width

            # 次の開始行（画像開始行 + 画像消費行数 + 空き行）
            self.excel_row = img_start_row + rows_consumed + gap
            
            self._save_excel_state(wb, ws.title)
            wb.save(xl_path)

            self._sync_ui_from_internal()
            self._update_excel_status()

            self._set_status(
                f"✓ Excel貼り付け完了 → {os.path.basename(xl_path)}"
                f"  [{ws.title}] {anchor}  次: 行 {self.excel_row}",
                SUCCESS,
            )
        except Exception as e:
            self._set_status(f"⚠ Excel貼り付けエラー: {e}", WARNING)

    def _start_hotkey(self):
        if sys.platform == "win32":
            hk = Win32GlobalHotkey(callback=self._on_hotkey_fired)
            if hk.start():
                self._win32_hotkey = hk
                self._hotkey_mode  = "win32"
                return

        def on_press(key):
            if key == pynput_keyboard.Key.f1 and self.f1_ready:
                self.f1_ready = False
                self._on_hotkey_fired()

        def on_release(key):
            if key == pynput_keyboard.Key.f1:
                self.f1_ready = True

        try:
            self._listener = pynput_keyboard.Listener(on_press=on_press, on_release=on_release)
            self._listener.daemon = True
            self._listener.start()
            self._hotkey_mode = "pynput"
        except Exception as e:
            self._hotkey_mode = "none"
            self._set_status(f"⚠ ホットキー登録失敗: {e}", WARNING)

    def _on_hotkey_fired(self):
        try:
            self.root.after(0, self._take_screenshot)
        except Exception:
            pass

    def _take_screenshot(self):
        if self._capturing:
            return
        if not self.region:
            self._set_status("⚠ 先に撮影範囲を指定してください", WARNING)
            return

        save_dir = self.save_dir.get()
        if not os.path.isdir(save_dir):
            try:
                os.makedirs(save_dir, exist_ok=True)
            except Exception as e:
                self._set_status(f"⚠ 保存先エラー: {e}", WARNING)
                return

        while True:
            filename = f"{self.counter:04d}.jpg"
            filepath = os.path.join(save_dir, filename)
            if not os.path.exists(filepath):
                break
            self.counter += 1

        self._capturing = True
        was_visible = False
        try:
            if self.root.state() == "normal" and self.root.focus_displayof() is not None:
                was_visible = True
                self.root.withdraw()
        except Exception:
            pass

        delay = 120 if was_visible else 0
        self.root.after(delay, lambda: self._do_capture(filepath, filename, was_visible))

    def _do_capture(self, filepath, filename, was_visible):
        try:
            # 1. ドラッグで取得した座標(論理ピクセル)
            x1, y1, x2, y2 = self.region
            
            # 2. Tkinterが認識しているメインモニタの解像度
            logical_w = self.root.winfo_screenwidth()
            logical_h = self.root.winfo_screenheight()

            # 3. Pillowでメインモニタ全体を撮影し、実際の物理解像度を取得
            full_img = ImageGrab.grab()
            physical_w, physical_h = full_img.size

            # 4. 解像度のズレを吸収する比率を計算
            scale_x = physical_w / logical_w
            scale_y = physical_h / logical_h

            # 5. Tkinterで取得した選択座標を、実際の画像ピクセル座標に補正
            rx1 = int(x1 * scale_x)
            ry1 = int(y1 * scale_y)
            rx2 = int(x2 * scale_x)
            ry2 = int(y2 * scale_y)

            # 6. 補正した座標で物理サイズの画像を切り抜き
            img = full_img.crop((rx1, ry1, rx2, ry2))

            # 7. 物理画像をドラッグ時の「見た目のサイズ(論理ピクセル)」にリサイズする
            logical_w_px = max(1, abs(x2 - x1))
            logical_h_px = max(1, abs(y2 - y1))
            
            if logical_w_px != img.width or logical_h_px != img.height:
                if hasattr(PILImage, "Resampling"):
                    resample_filter = PILImage.Resampling.LANCZOS
                else:
                    resample_filter = PILImage.LANCZOS
                img = img.resize((logical_w_px, logical_h_px), resample_filter)

            # 【重要】Excel側で勝手に画像が巨大化しないように、DPIを96に強制固定して保存する
            img.save(filepath, "JPEG", quality=95, dpi=(96, 96))
            
            self._set_status(f"✓ 保存完了: {filename}  [{img.width}×{img.height}px]", SUCCESS)
            self.counter += 1
            self._show_flash()

            # リサイズ済みの物理的な幅・高さを渡す（これが論理サイズと同一になっている）
            self._paste_to_excel(filepath, img.width, img.height)

        except Exception as e:
            self._set_status(f"⚠ 撮影エラー: {e}", WARNING)
        finally:
            if was_visible:
                self.root.deiconify()
            self._capturing = False

    def _show_flash(self):
        self.status_bar.config(bg=ACCENT)
        self.root.after(150, lambda: self.status_bar.config(bg=PANEL))

    def _sync_counter(self):
        d = self.save_dir.get()
        if not os.path.isdir(d):
            self.counter = 1
            return
        existing = [
            int(os.path.splitext(f)[0])
            for f in os.listdir(d)
            if os.path.splitext(f)[1].lower() == ".jpg" and os.path.splitext(f)[0].isdigit()
        ]
        self.counter = (max(existing) + 1) if existing else 1

    def _set_status(self, msg, color=SUBTEXT):
        self.status_var.set(msg)
        self.status_bar.config(fg=color)

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        if self._hotkey_mode == "win32":
            self._set_status("準備完了（F1グローバル登録: Win32 API）", SUCCESS)
        elif self._hotkey_mode == "pynput":
            self._set_status("準備完了（F1グローバル登録: pynput）", SUCCESS)
        else:
            self._set_status("⚠ F1ホットキーが登録できませんでした", WARNING)
        self.root.mainloop()

    def _on_close(self):
        if self._win32_hotkey:
            try:
                self._win32_hotkey.stop()
            except Exception:
                pass
            self._win32_hotkey = None
        if self._listener:
            try:
                self._listener.stop()
            except Exception:
                pass
            self._listener = None
        self.root.destroy()


if __name__ == "__main__":
    app = ScreenshotApp()
    app.run()